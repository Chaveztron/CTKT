from django.shortcuts import render, redirect, reverse, get_object_or_404
from django.template.loader import get_template

from .forms import *
from django.contrib import messages
from openpyxl import Workbook
from django.http.response import HttpResponse
from django.views.generic import TemplateView, View

from .utils import render_to_pdf #created in step 4

from django.core.mail import send_mail

from django.contrib.auth.decorators import login_required

from django.views.generic import TemplateView

from .models import *

from django.db.models import Count

class HomeView(TemplateView):
    template_name = "home.html"

def comming(request):
    return render(request, 'evento/index2.html')

def login(request):
    return render(request, 'evento/login.html')

def programa(request):
    compradores = Comprador.objects.all()
    videos = Video.objects.all()
    exposiciones = Publicaciones.objects.all()
    autores_relevantes = AutoresRelevantes.objects.all()
    texto_adicional = TextoAdicional.objects.all()
    context = {
        'videos': videos,
        'compradores': compradores,
        'Exposiciones': exposiciones,
        'Autores': autores_relevantes,
        'Texto': texto_adicional,
    }
    return render(request, 'evento/programa.html', context)

# Create your views here.
def index(request):
    compradores = len(Usuario.objects.extra(where=["participaciones_id = 1"]))
    proveedores = len(Usuario.objects.extra(where=["participaciones_id = 2"]))
    visitantes = len(Usuario.objects.extra(where=["participaciones_id = 3"]))
    form = RegistrerForm(request.POST)
    if request.method == "POST":
        if form.is_valid():
            form.save()
            messages.success(request, f'Registro exitoso')
            return redirect(reverse("registrado", kwargs={
                'id': form.instance.id,
                'nombre': form.instance.nombre
            }))
    else:
        form = RegistrerForm()
    context = {
        'form':form,
        'c': compradores,
        'p': proveedores,
        'v': visitantes,
    }
    return render(request, 'evento/index.html', context)

def indexen(request):
    formen = FormEn(request.POST)
    if request.method == "POST":
        if formen.is_valid():
            formen.save()
            messages.success(request, f'Registro exitoso')
            return redirect(reverse("registrado", kwargs={
                'id': formen.instance.id,
                'nombre': formen.instance.nombre
            }))
    else:
        formen = FormEn()
    context = {
        'form':formen
    }
    return render(request, 'evento/en.html', context)

def satisfatorio(request, id, nombre):
    #user = get_object_or_404(Usuario, id=id)
    dominio = str(dns.objects.first())
    user = Usuario.objects.get(id = id, nombre = nombre)
    nombreUser = user.nombre
    nombreUser = nombreUser.replace(" ", "%20")
    link = "http://"+dominio+"/pase/" + str(user.id) + "/" + nombreUser + "/" ## cambiar dominio

    send_mail('Hola desde EXPO INDUSTRIA TECATE',
              'Usted a generado un boleto de entrada,'
              'puede descargarlo desde aqui ' + link,#########
              'kyworkred@gmail.com',
              [user.email],
              fail_silently=False)
    context = {
        "id": user.id,
        "nombre": user.nombre,
    }
    return render(request, 'evento/satisfactorio.html', context)

@login_required
def verificar(request, id, nombre):
    user = Usuario.objects.get(id=id, nombre=nombre)
    user.asistencia = True
    user.save()
    name = str(user.nombre) + " " + str(user.appellidoP)
    '''persona = Saludo(persona=str(user.nombre))
    persona.save()'''

    usuario = Usuario.objects.get(id=id, nombre=nombre)
    nombreUser = usuario.nombre
    nombreUser = nombreUser.replace(" ", "%20")
    link = "BEGIN:VCARD\n" \
           "VERSION:3.0\n" \
           "N:"+str(usuario.nombre)+";"+str(usuario.appellidoP)+";;Mr.;\n" \
           "FN:"+str(usuario.nombre)+" "+str(usuario.appellidoP)+"\n" \
           "ORG:"+str(usuario.empresa)+".\n" \
           "TITLE:"+str(usuario.puesto)+"\n" \
            "TEL;TYPE=WORK,VOICE:" + str(usuario.telefono) + "\n" \
            "EMAIL:"+str(usuario.email)+"\n" \
           "END:VCARD\n"
    template = get_template('pdf/invoice.html')
    context = {
        "name":name,
        "nombre": usuario.nombre.upper(),
        "apellidoP": usuario.appellidoP.upper(),
        "puesto": usuario.puesto,
        "empresa": usuario.empresa.upper(),
        "email": usuario.email,
        "telefono": usuario.telefono,
        "participacion": usuario.participaciones.tipo_parti,
        "hora": usuario.horaRegistro,
        "link": link,
    }
    html = template.render(context)
    pdf = render_to_pdf('pdf/sticker.html', context)
    if pdf:
        response = HttpResponse(pdf, content_type='application/pdf')
        filename = "Invoice_%s.pdf" % ("12341231")
        content = "inline; filename='%s'" % (filename)
        download = request.GET.get("download")
        if download:
            content = "attachment; filename='%s'" % (filename)
        response['Content-Disposition'] = content
        # http://localhost:8000/pase/1/?download=true
        return response
    return HttpResponse("Not found")

@login_required
def registrados(request):
    registrados = Usuario.objects.order_by('-horaRegistro')

    compradores = len(Usuario.objects.extra(where=["participaciones_id = 1"]))
    proveedores = len(Usuario.objects.extra(where=["participaciones_id = 2"]))
    visitantes = len(Usuario.objects.extra(where=["participaciones_id = 3"]))



    context = {
        'usuarios': registrados,
        'c': compradores,
        'p': proveedores,
        'v': visitantes,
    }
    return render(request, 'evento/tabulado.html', context)

class ReporteUsuarioExcel(TemplateView):
    def get(self, request, *args, **kwargs):
        usuarios = Usuario.objects.order_by('-horaRegistro')
        wb = Workbook()
        ws = wb.active
        ws['B1'] = 'Lista de asistencia'

        ws.merge_cells('B1:E1')

        ws['A3'] = 'ID'
        ws['B3'] = 'Nombres'
        ws['C3'] = 'Puestos'
        ws['D3'] = 'Empresas'
        ws['E3'] = 'Email'
        ws['F3'] = 'Telefonos'
        ws['G3'] = 'Intereses'
        ws['H3'] = 'Participacion'
        ws['I3'] = 'Hora de registro'
        ws['J3'] = 'Asistencia'

        cont = 4

        for usuario in usuarios:
            ws.cell(row = cont, column = 1).value = usuario.id
            ws.cell(row = cont, column = 2).value = str(usuario.nombre)+" "+ str(usuario.appellidoP)
            ws.cell(row = cont, column = 3).value = usuario.puesto
            ws.cell(row = cont, column = 4).value = usuario.empresa
            ws.cell(row = cont, column = 5).value = usuario.email
            ws.cell(row = cont, column = 6).value = usuario.telefono

            intereses = ""
            for int in usuario.intereses.all():
                intereses = str(int) +", "+intereses

            ws.cell(row = cont, column = 7).value = intereses

            ws.cell(row = cont, column = 8).value = usuario.participaciones.tipo_parti
            ws.cell(row = cont, column = 9).value = usuario.horaRegistro
            ws.cell(row = cont, column = 10).value = usuario.asistencia
            cont+=1

        nombre_archivo = "ReporteUsuariosRegistrados.xlsx"
        response = HttpResponse(content_type="application/ms-excel")
        content = "attachment; filename = {0}".format(nombre_archivo)
        response['Content-Disposition'] = content
        wb.save(response)
        return response

def GeneratePdf(request,id, nombre):
    #usuario = get_object_or_404(Usuario, nombre=nombre)
    dominio = str(dns.objects.first())
    usuario = Usuario.objects.get(id = id, nombre = nombre)
    nombreUser = usuario.nombre
    nombreUser = nombreUser.replace(" ","%20")
    link = "http://"+dominio+"/verificado/"+str(usuario.id)+"/"+nombreUser+"/"  ##cambiar dominio
    template = get_template('pdf/invoice.html')
    name = str(usuario.nombre)+" "+str(usuario.appellidoP)
    context = {
        "nombre": name,
        "puesto": usuario.puesto,
        "empresa": usuario.empresa,
        "email": usuario.email,
        "telefono": usuario.telefono,
        "participacion": usuario.participaciones.tipo_parti,
        "hora": usuario.horaRegistro,
        "link": link,
    }
    html = template.render(context)
    pdf = render_to_pdf('pdf/invoice.html', context)
    if pdf:
        response = HttpResponse(pdf, content_type='application/pdf')
        filename = "Invoice_%s.pdf" % ("12341231")
        content = "inline; filename='%s'" % (filename)
        download = request.GET.get("download")
        if download:
            content = "attachment; filename='%s'" % (filename)
        response['Content-Disposition'] = content
        #http://localhost:8000/pase/1/?download=true
        return response
    return HttpResponse("Not found")